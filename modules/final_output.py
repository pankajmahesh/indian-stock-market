"""
Step 7: Final Top 20 Output
Re-rank top 50 using composite + qualitative scores.
Generate actionable report with thesis, risks, entry zones, stop-loss.
"""
import os

import numpy as np
import pandas as pd

import config
from modules.data_fetcher import DataFetcher
from utils.helpers import format_indian_crores, format_pct
from utils.logger import log


class FinalOutput:
    def __init__(self, data_fetcher: DataFetcher):
        self.fetcher = data_fetcher

    def generate(self, deep_dive_df):
        """
        Re-rank top 50, select top 20, and generate final report.
        """
        log.info("=" * 60)
        log.info("STEP 7: FINAL TOP 20 OUTPUT")
        log.info("=" * 60)

        df = deep_dive_df.copy()

        # Final score = 70% composite + 30% qualitative
        df["final_score"] = (
            config.FINAL_COMPOSITE_WEIGHT * df["composite_score"]
            + config.FINAL_QUALITATIVE_WEIGHT * df["qualitative_score"].fillna(0)
        )

        # Sort and pick top 20
        df = df.sort_values("final_score", ascending=False).reset_index(drop=True)
        top20 = df.head(config.TOP_N_FINAL).copy()

        # Enrich with entry zones, stop-loss, thesis, risks
        symbols = top20["symbol"].tolist()
        all_prices = self.fetcher.batch_download_prices(symbols)
        all_info = self.fetcher.batch_fetch_info(symbols)

        enriched = []
        for idx, row in top20.iterrows():
            sym = row["symbol"]
            prices = all_prices.get(sym)
            info = all_info.get(sym, {})

            entry_zone = self._compute_entry_zone(prices, row.get("last_price"))
            stop_loss = self._compute_stop_loss(prices, row.get("last_price"))
            target = self._compute_target(info, row.get("last_price"))
            thesis = self._generate_thesis(row, info)
            risk = self._generate_risk(row, info)

            # Price change %
            cmp_val = row.get("last_price")
            prev_close = info.get("previousClose") or info.get("regularMarketPreviousClose")
            change_pct = None
            if cmp_val and prev_close and prev_close > 0:
                change_pct = round((cmp_val - prev_close) / prev_close * 100, 2)

            # Promoter / insider holding
            promoter_holding = info.get("heldPercentInsiders")
            if promoter_holding is not None:
                promoter_holding = round(promoter_holding * 100, 1)

            enriched.append({
                "rank": idx + 1,
                "symbol": sym,
                "name": row.get("name", ""),
                "sector": row.get("sector", ""),
                "industry": row.get("industry", ""),
                "l_category": row.get("l_category", "L2"),
                "cmp": cmp_val,
                "change_pct": change_pct,
                "market_cap_cr": round(row.get("market_cap", 0) / 1e7, 1),
                "promoter_holding_pct": promoter_holding,
                "final_score": round(row["final_score"], 1),
                "fundamental_score": round(row.get("fundamental_score", 0), 1),
                "technical_score": round(row.get("technical_score", 0), 1),
                "qualitative_score": round(row.get("qualitative_score", 0), 1),
                "composite_score": round(row.get("composite_score", 0), 1),
                "qual_strategy_alignment": round(row.get("qual_strategy_alignment", 0) or 0, 1),
                "bull_thesis": thesis,
                "key_risk": risk,
                "entry_zone": entry_zone,
                "stop_loss": stop_loss,
                "target": target,
                "pe_ratio": row.get("pe_ratio"),
                "roe": row.get("roe"),
                "debt_to_equity": row.get("debt_to_equity"),
                "data_quality": row.get("data_quality", ""),
            })

        result_df = pd.DataFrame(enriched)

        # Save CSV
        csv_path = os.path.join(config.OUTPUT_DIR, "final_top20.csv")
        result_df.to_csv(csv_path, index=False)
        log.info(f"Saved CSV: {csv_path}")

        # Save Excel with formatting
        self._save_excel(result_df)

        # Print console report
        self._print_report(result_df)

        return result_df

    # ----------------------------------------------------------
    # Entry zone computation
    # ----------------------------------------------------------
    def _compute_entry_zone(self, prices, cmp):
        if prices is None or prices.empty or cmp is None:
            return "N/A"
        try:
            close = prices["Close"].astype(float)

            # EMA50 as support
            ema50 = close.ewm(span=50, adjust=False).mean().iloc[-1]

            # Recent swing low (last 20 days)
            recent_low = close.tail(20).min()

            support = max(ema50, recent_low)
            entry_high = cmp * 0.97  # 3% below CMP

            low = round(support, 1)
            high = round(entry_high, 1)

            if low > high:
                low, high = high, low

            return f"{low} - {high}"
        except Exception:
            return "N/A"

    # ----------------------------------------------------------
    # Stop-loss computation
    # ----------------------------------------------------------
    def _compute_stop_loss(self, prices, cmp):
        if prices is None or prices.empty or cmp is None:
            return "N/A"
        try:
            close = prices["Close"].astype(float)

            # EMA200 or -15% from CMP, whichever is tighter
            if len(close) >= 200:
                ema200 = close.ewm(span=200, adjust=False).mean().iloc[-1]
            else:
                ema200 = cmp * 0.85

            pct_stop = cmp * 0.85
            stop = min(ema200, pct_stop)

            return round(stop, 1)
        except Exception:
            return "N/A"

    # ----------------------------------------------------------
    # Target computation
    # ----------------------------------------------------------
    def _compute_target(self, info, cmp):
        target = info.get("targetMeanPrice")
        if target is not None:
            return round(target, 1)
        if cmp is not None:
            return round(cmp * 1.2, 1)  # 20% upside as fallback
        return "N/A"

    # ----------------------------------------------------------
    # Auto-generated thesis
    # ----------------------------------------------------------
    def _generate_thesis(self, row, info):
        parts = []

        # Strongest fundamental category
        fund_cats = [
            ("Profitability", row.get("fund_profitability")),
            ("Growth", row.get("fund_growth")),
            ("Valuation", row.get("fund_valuation")),
            ("Financial Health", row.get("fund_financial_health")),
            ("Dividend", row.get("fund_dividend")),
        ]
        fund_cats = [(n, s) for n, s in fund_cats if s is not None and s > 0]
        fund_cats.sort(key=lambda x: x[1], reverse=True)

        if fund_cats:
            top = fund_cats[0]
            parts.append(f"Strong {top[0]} ({top[1]:.0f}/100)")

        # Technical trend
        tech_trend = row.get("tech_trend")
        if tech_trend is not None and tech_trend > 60:
            parts.append("bullish technical trend")
        elif tech_trend is not None and tech_trend > 40:
            parts.append("neutral-to-positive technicals")

        # Analyst target
        target = info.get("targetMeanPrice")
        cmp = row.get("last_price")
        if target and cmp and cmp > 0:
            upside = ((target - cmp) / cmp) * 100
            if upside > 10:
                parts.append(f"{upside:.0f}% analyst upside")

        # Sector and L-category
        sector = row.get("sector", "")
        l_cat = row.get("l_category", "")
        if sector and l_cat:
            parts.append(f"Sector: {sector} ({l_cat})")
        elif sector:
            parts.append(f"Sector: {sector}")

        # Strategy alignment
        sa_score = row.get("qual_strategy_alignment")
        if sa_score is not None and sa_score > 70:
            parts.append("strong Bandhan strategy alignment")

        return ". ".join(parts) if parts else "Data insufficient for thesis generation"

    # ----------------------------------------------------------
    # Auto-generated risk
    # ----------------------------------------------------------
    def _generate_risk(self, row, info):
        risks = []

        # L3 cyclicality warning
        l_cat = row.get("l_category", "")
        if l_cat == "L3":
            risks.append("L3 cyclical stock — earnings volatile across cycles")

        # Valuation concern
        pe = row.get("pe_ratio")
        if pe is not None and pe > 40:
            risks.append(f"Expensive valuation (PE: {pe:.1f})")

        # Debt concern
        de = row.get("debt_to_equity")
        if de is not None and de > 100:
            risks.append(f"Elevated debt (D/E: {de:.0f}%)")

        # Cautious sector warning (IT services automation risk)
        sector = row.get("sector", "")
        if sector == "Technology":
            risks.append("IT sector faces automation disruption risk (Bandhan: underweight)")
        elif sector == "Energy":
            risks.append("Oil/Gas sector — linear growth, cautious per Bandhan thesis")
        elif sector == "Utilities":
            risks.append("Utility — linear GDP-linked growth, limited alpha potential")

        # Low data quality
        dq = row.get("data_quality", "")
        if dq == "LOW":
            risks.append("Limited data availability")

        # Weak technicals
        tech = row.get("technical_score", 0)
        if tech < 30:
            risks.append("Weak technical momentum")

        # Sector concentration
        if not risks:
            risks.append("Sector-level macro risks")

        return ". ".join(risks[:2])

    # ----------------------------------------------------------
    # Excel export with formatting
    # ----------------------------------------------------------
    def _save_excel(self, df):
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter

            excel_path = os.path.join(config.OUTPUT_DIR, "final_top20_report.xlsx")
            df.to_excel(excel_path, index=False, sheet_name="Top 20 Stocks")

            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            ws = wb.active

            # Header styling
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Score color coding
            score_cols = []
            for idx, col in enumerate(df.columns, 1):
                if "score" in col.lower():
                    score_cols.append(idx)

            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            for row_idx in range(2, len(df) + 2):
                for col_idx in score_cols:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        val = float(cell.value) if cell.value else 0
                        if val >= 70:
                            cell.fill = green
                        elif val >= 50:
                            cell.fill = yellow
                        else:
                            cell.fill = red
                    except (ValueError, TypeError):
                        pass

            # Auto-fit column widths
            for col_idx in range(1, len(df.columns) + 1):
                letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(1, len(df) + 2)
                )
                ws.column_dimensions[letter].width = min(max_len + 2, 40)

            wb.save(excel_path)
            log.info(f"Saved Excel: {excel_path}")

        except Exception as e:
            log.warning(f"Excel formatting failed (CSV still saved): {e}")

    # ----------------------------------------------------------
    # Console report
    # ----------------------------------------------------------
    def _print_report(self, df):
        print("\n" + "=" * 100)
        print("  INDIAN STOCK SCREENER — FINAL TOP 20 PICKS  (Bandhan AMC Strategy)")
        print("=" * 100)

        for _, row in df.iterrows():
            l_cat = row.get("l_category", "")
            l_label = f"  [{l_cat}]" if l_cat else ""
            print(f"\n{'─' * 80}")
            print(f"  #{row['rank']}  {row['symbol']}{l_label}  —  {row['name']}")
            print(f"  Sector: {row['sector']}  |  Industry: {row['industry']}")
            print(f"  CMP: ₹{row['cmp']:,.1f}  |  Market Cap: ₹{row['market_cap_cr']:,.0f} Cr")
            print(f"  ┌─ Scores ────────────────────────────────────────────────────")
            print(f"  │  Final: {row['final_score']:.1f}  |  Fundamental: {row['fundamental_score']:.1f}  |  Technical: {row['technical_score']:.1f}  |  Qualitative: {row['qualitative_score']:.1f}  |  Strategy: {row['qual_strategy_alignment']:.1f}")
            print(f"  ├─ Thesis ───────────────────────────────────────────────────")
            print(f"  │  {row['bull_thesis']}")
            print(f"  ├─ Risk ─────────────────────────────────────────────────────")
            print(f"  │  {row['key_risk']}")
            print(f"  ├─ Levels ───────────────────────────────────────────────────")
            print(f"  │  Entry Zone: ₹{row['entry_zone']}  |  Stop Loss: ₹{row['stop_loss']}  |  Target: ₹{row['target']}")
            print(f"  └────────────────────────────────────────────────────────────")

        # Portfolio balance summary (Bandhan: avoid over-concentration in L1 or L3)
        print(f"\n{'─' * 100}")
        print("  PORTFOLIO BALANCE (Bandhan AMC Strategy)")
        print(f"{'─' * 100}")
        if "l_category" in df.columns:
            counts = df["l_category"].value_counts().to_dict()
            total = len(df)
            l1 = counts.get("L1", 0)
            l2 = counts.get("L2", 0)
            l3 = counts.get("L3", 0)
            print(f"  L1 (High Quality):  {l1:2d} stocks  ({l1/total*100:.0f}%)")
            print(f"  L2 (Mid Quality):   {l2:2d} stocks  ({l2/total*100:.0f}%)")
            print(f"  L3 (Cyclical):      {l3:2d} stocks  ({l3/total*100:.0f}%)")

            # Balance warnings
            if l1 / total > 0.70:
                print("  ⚠  Heavy L1 concentration — consider adding selective cyclical exposure")
            if l3 / total > 0.40:
                print("  ⚠  High L3 cyclical exposure — ensure cycle timing is right")
            if l2 / total >= 0.40:
                print("  ✓  Good L2 balance — broad stock-selection alpha base")

        print(f"\n{'=' * 100}")
        print(f"  Report generated. Files saved to: {config.OUTPUT_DIR}")
        print(f"{'=' * 100}\n")

    @staticmethod
    def load_saved():
        path = os.path.join(config.OUTPUT_DIR, "final_top20.csv")
        if os.path.exists(path):
            return pd.read_csv(path)
        return None
